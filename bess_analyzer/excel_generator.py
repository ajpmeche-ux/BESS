"""Generate Excel-based BESS Economic Analysis Workbook.

Creates a macro-enabled Excel workbook (.xlsm) with:
- Project input forms
- Assumption library selection (NREL, Lazard, CPUC)
- Automatic calculation engine using Excel formulas
- Results dashboard with conditional formatting
- Annual cash flow projections
- Methodology documentation with formulas and citations
- VBA macros for report generation

Usage:
    python excel_generator.py [output_path]
"""

import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule, ColorScaleRule
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from openpyxl.workbook.defined_name import DefinedName


def create_styles():
    """Create reusable cell styles."""
    styles = {}

    # Header style
    styles['header'] = NamedStyle(name='header')
    styles['header'].font = Font(bold=True, color='FFFFFF', size=11)
    styles['header'].fill = PatternFill('solid', fgColor='1565C0')
    styles['header'].alignment = Alignment(horizontal='center', vertical='center')

    # Section header
    styles['section'] = NamedStyle(name='section')
    styles['section'].font = Font(bold=True, size=12, color='1565C0')
    styles['section'].fill = PatternFill('solid', fgColor='E3F2FD')

    # Input cell
    styles['input'] = NamedStyle(name='input')
    styles['input'].fill = PatternFill('solid', fgColor='FFFDE7')
    styles['input'].border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Formula cell
    styles['formula'] = NamedStyle(name='formula')
    styles['formula'].fill = PatternFill('solid', fgColor='E8F5E9')
    styles['formula'].border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Result cell
    styles['result'] = NamedStyle(name='result')
    styles['result'].font = Font(bold=True, size=14)
    styles['result'].fill = PatternFill('solid', fgColor='E3F2FD')
    styles['result'].alignment = Alignment(horizontal='center')

    # Currency format
    styles['currency'] = NamedStyle(name='currency')
    styles['currency'].number_format = '$#,##0'

    # Percentage format
    styles['percent'] = NamedStyle(name='percent')
    styles['percent'].number_format = '0.0%'

    return styles


def create_inputs_sheet(wb, styles):
    """Create the Project Inputs sheet."""
    ws = wb.active
    ws.title = "Inputs"

    # Column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 40

    row = 2

    # Title
    ws.merge_cells('B2:E2')
    ws['B2'] = "BESS Economic Analysis - Project Inputs"
    ws['B2'].font = Font(bold=True, size=16, color='1565C0')
    row = 4

    # Section: Project Basics
    ws[f'B{row}'] = "PROJECT BASICS"
    ws[f'B{row}'].style = styles['section']
    ws.merge_cells(f'B{row}:E{row}')
    row += 1

    inputs = [
        ('Project Name', 'Inputs!C5', '', 'Text', 'Enter project name'),
        ('Project ID', 'Inputs!C6', '', 'Text', 'Unique identifier'),
        ('Location', 'Inputs!C7', '', 'Text', 'Site or market location (e.g., CAISO NP15)'),
        ('Capacity (MW)', 'Inputs!C8', 100, 'Number', 'Nameplate power capacity'),
        ('Duration (hours)', 'Inputs!C9', 4, 'Number', 'Storage duration'),
        ('Energy Capacity (MWh)', 'Inputs!C10', '=C8*C9', 'Formula', 'Auto-calculated: MW × hours'),
        ('Analysis Period (years)', 'Inputs!C11', 20, 'Number', 'Economic analysis horizon'),
        ('Discount Rate (%)', 'Inputs!C12', 7.0, 'Percent', 'Nominal discount rate for NPV'),
        ('Ownership Type', 'Inputs!C13', 'Utility', 'Text', 'Utility (lower WACC ~6-7%) or Merchant (~8-10%)'),
    ]

    for label, cell_ref, value, dtype, tooltip in inputs:
        ws[f'B{row}'] = label
        ws[f'B{row}'].font = Font(bold=True)
        cell = ws[f'C{row}']
        if dtype == 'Formula':
            cell.value = value
            cell.style = styles['formula']
        else:
            cell.value = value
            cell.style = styles['input']
        if dtype == 'Percent':
            cell.number_format = '0.0%'
            cell.value = value / 100
        ws[f'D{row}'] = 'MW' if 'MW' in label else ('hrs' if 'hour' in label else ('yrs' if 'year' in label else ''))
        ws[f'E{row}'] = tooltip
        ws[f'E{row}'].font = Font(italic=True, color='666666')
        row += 1

    row += 1

    # Section: Assumption Library Selection
    ws[f'B{row}'] = "ASSUMPTION LIBRARY"
    ws[f'B{row}'].style = styles['section']
    ws.merge_cells(f'B{row}:E{row}')
    row += 1

    ws[f'B{row}'] = "Select Library"
    ws[f'B{row}'].font = Font(bold=True)
    library_row = row
    ws[f'C{row}'] = "NREL ATB 2024 - Moderate"
    ws[f'C{row}'].style = styles['input']

    # Add dropdown for library selection
    dv = DataValidation(
        type="list",
        formula1='"NREL ATB 2024 - Moderate,Lazard LCOS 2025,CPUC California 2024,Custom"',
        allow_blank=False
    )
    dv.error = "Please select a valid library"
    dv.prompt = "Choose an assumption library"
    ws.add_data_validation(dv)
    dv.add(f'C{library_row}')

    ws[f'E{row}'] = "Select library to auto-populate costs and benefits"
    ws[f'E{row}'].font = Font(italic=True, color='666666')
    row += 2

    # Section: Technology Specs
    ws[f'B{row}'] = "TECHNOLOGY SPECIFICATIONS"
    ws[f'B{row}'].style = styles['section']
    ws.merge_cells(f'B{row}:E{row}')
    row += 1

    tech_inputs = [
        ('Chemistry', 'LFP', 'Text', 'LFP, NMC, or Other'),
        ('Round-Trip Efficiency (%)', 0.85, 'Percent', 'AC-AC efficiency'),
        ('Annual Degradation (%)', 0.025, 'Percent', 'Capacity loss per year'),
        ('Cycle Life', 6000, 'Number', 'Full-depth cycles before EOL'),
        ('Augmentation Year', 12, 'Number', 'Year of battery replacement'),
    ]

    tech_start = row
    for i, (label, value, dtype, tooltip) in enumerate(tech_inputs):
        ws[f'B{row}'] = label
        ws[f'B{row}'].font = Font(bold=True)
        cell = ws[f'C{row}']
        cell.value = value
        cell.style = styles['input']
        if dtype == 'Percent':
            cell.number_format = '0.0%'
        ws[f'E{row}'] = tooltip
        ws[f'E{row}'].font = Font(italic=True, color='666666')
        row += 1

    row += 1

    # Section: Cost Inputs
    ws[f'B{row}'] = "COST INPUTS"
    ws[f'B{row}'].style = styles['section']
    ws.merge_cells(f'B{row}:E{row}')
    row += 1

    cost_inputs = [
        ('CapEx ($/kWh)', 160, '$/kWh', 'Installed capital cost per kWh'),
        ('Fixed O&M ($/kW-year)', 25, '$/kW-yr', 'Annual fixed O&M'),
        ('Variable O&M ($/MWh)', 0, '$/MWh', 'Per-MWh discharge cost'),
        ('Augmentation Cost ($/kWh)', 55, '$/kWh', 'Battery replacement cost'),
        ('Decommissioning ($/kW)', 10, '$/kW', 'End-of-life cost'),
    ]

    cost_start = row
    for i, (label, value, unit, tooltip) in enumerate(cost_inputs):
        ws[f'B{row}'] = label
        ws[f'B{row}'].font = Font(bold=True)
        cell = ws[f'C{row}']
        cell.value = value
        cell.style = styles['input']
        cell.number_format = '$#,##0.00'
        ws[f'D{row}'] = unit
        ws[f'E{row}'] = tooltip
        ws[f'E{row}'].font = Font(italic=True, color='666666')
        row += 1

    row += 1

    # Section: Tax Credits (BESS-Specific)
    ws[f'B{row}'] = "TAX CREDITS (BESS-Specific under IRA)"
    ws[f'B{row}'].style = styles['section']
    ws.merge_cells(f'B{row}:E{row}')
    row += 1

    tax_inputs = [
        ('ITC Base Rate (%)', 0.30, 'Percent', '30% Investment Tax Credit under IRA'),
        ('ITC Adders (%)', 0.0, 'Percent', 'Energy community +10%, Domestic content +10%'),
    ]

    for label, value, dtype, tooltip in tax_inputs:
        ws[f'B{row}'] = label
        ws[f'B{row}'].font = Font(bold=True)
        cell = ws[f'C{row}']
        cell.value = value
        cell.style = styles['input']
        if dtype == 'Percent':
            cell.number_format = '0.0%'
        ws[f'E{row}'] = tooltip
        ws[f'E{row}'].font = Font(italic=True, color='666666')
        row += 1

    row += 1

    # Section: Infrastructure Costs (Common to all utility projects)
    ws[f'B{row}'] = "INFRASTRUCTURE COSTS (Common to all projects)"
    ws[f'B{row}'].style = styles['section']
    ws.merge_cells(f'B{row}:E{row}')
    row += 1

    infra_inputs = [
        ('Interconnection ($/kW)', 100, '$/kW', 'Network upgrades, studies'),
        ('Land ($/kW)', 10, '$/kW', 'Site acquisition/lease'),
        ('Permitting ($/kW)', 15, '$/kW', 'Permits, environmental review'),
        ('Insurance (% of CapEx)', 0.005, 'Percent', 'Annual insurance cost'),
        ('Property Tax (%)', 0.01, 'Percent', 'Annual property tax rate'),
    ]

    for label, value, dtype, tooltip in infra_inputs:
        ws[f'B{row}'] = label
        ws[f'B{row}'].font = Font(bold=True)
        cell = ws[f'C{row}']
        cell.value = value
        cell.style = styles['input']
        if dtype == 'Percent':
            cell.number_format = '0.00%'
        else:
            cell.number_format = '$#,##0.00'
        ws[f'E{row}'] = tooltip
        ws[f'E{row}'].font = Font(italic=True, color='666666')
        row += 1

    row += 1

    # Section: Benefit Inputs
    ws[f'B{row}'] = "BENEFIT STREAMS (Year 1 Values)"
    ws[f'B{row}'].style = styles['section']
    ws.merge_cells(f'B{row}:E{row}')
    row += 1

    # Headers for benefits table
    ws[f'B{row}'] = "Benefit Category"
    ws[f'C{row}'] = "$/kW-year"
    ws[f'D{row}'] = "Escalation %"
    ws[f'E{row}'] = "Citation"
    for col in ['B', 'C', 'D', 'E']:
        ws[f'{col}{row}'].style = styles['header']
    row += 1

    benefits_start = row
    benefits = [
        ('Resource Adequacy', 150, 0.02, 'CPUC RA Report 2024', 'common'),
        ('Energy Arbitrage', 40, 0.015, 'CAISO Market Data 2024', 'common'),
        ('Ancillary Services', 15, 0.01, 'CAISO AS Reports 2024', 'common'),
        ('T&D Deferral', 25, 0.02, 'CPUC Avoided Cost Calculator 2024', 'common'),
        ('Resilience Value', 50, 0.02, 'LBNL ICE Calculator 2024', 'common'),
        ('Renewable Integration', 25, 0.02, 'NREL Grid Integration Studies', 'bess_specific'),
        ('GHG Emissions Value', 15, 0.03, 'EPA Social Cost of Carbon 2024', 'bess_specific'),
        ('Voltage Support', 8, 0.01, 'EPRI Distribution Studies', 'common'),
    ]

    for name, value, esc, cite, category in benefits:
        ws[f'B{row}'] = name
        ws[f'C{row}'] = value
        ws[f'C{row}'].style = styles['input']
        ws[f'C{row}'].number_format = '$#,##0'
        ws[f'D{row}'] = esc
        ws[f'D{row}'].style = styles['input']
        ws[f'D{row}'].number_format = '0.0%'
        # Add category indicator with citation
        ws[f'E{row}'] = f"[{category}] {cite}"
        row += 1

    row += 1

    # Section: Cost Projections (Learning Curve)
    ws[f'B{row}'] = "COST PROJECTIONS (Learning Curve)"
    ws[f'B{row}'].style = styles['section']
    ws.merge_cells(f'B{row}:E{row}')
    row += 1

    cost_proj_inputs = [
        ('Annual Cost Decline Rate', 0.10, 'Percent', 'Technology learning rate (10-15% typical for batteries)'),
        ('Cost Base Year', 2024, 'Number', 'Reference year for base costs'),
    ]

    learning_rate_row = row
    for label, value, dtype, tooltip in cost_proj_inputs:
        ws[f'B{row}'] = label
        ws[f'B{row}'].font = Font(bold=True)
        cell = ws[f'C{row}']
        cell.value = value
        cell.style = styles['input']
        if dtype == 'Percent':
            cell.number_format = '0.0%'
        ws[f'E{row}'] = tooltip
        ws[f'E{row}'].font = Font(italic=True, color='666666')
        row += 1

    row += 1

    # Future Cost Analysis Section
    ws[f'B{row}'] = "FUTURE COST PROJECTIONS"
    ws[f'B{row}'].style = styles['section']
    ws.merge_cells(f'B{row}:E{row}')
    row += 1

    # Show projected costs at key years
    ws[f'B{row}'] = "Year"
    ws[f'C{row}'] = "Projected CapEx ($/kWh)"
    ws[f'D{row}'] = "Augmentation Cost ($/kWh)"
    ws[f'E{row}'] = "% of Today's Cost"
    for col in ['B', 'C', 'D', 'E']:
        ws[f'{col}{row}'].style = styles['header']
    row += 1

    projection_start = row
    for offset in [0, 5, 10, 12, 15, 20]:
        ws[f'B{row}'] = f'=C{learning_rate_row+1}+{offset}'
        # CapEx projection with learning curve
        ws[f'C{row}'] = f'=C24*(1-$C${learning_rate_row})^{offset}'
        ws[f'C{row}'].style = styles['formula']
        ws[f'C{row}'].number_format = '$#,##0.00'
        # Augmentation cost projection
        ws[f'D{row}'] = f'=C27*(1-$C${learning_rate_row})^{offset}'
        ws[f'D{row}'].style = styles['formula']
        ws[f'D{row}'].number_format = '$#,##0.00'
        # Percentage
        ws[f'E{row}'] = f'=(1-$C${learning_rate_row})^{offset}'
        ws[f'E{row}'].style = styles['formula']
        ws[f'E{row}'].number_format = '0.0%'
        row += 1

    row += 1

    # Define named ranges for easy reference
    wb.defined_names['Capacity_MW'] = DefinedName('Capacity_MW', attr_text='Inputs!$C$8')
    wb.defined_names['Duration_Hours'] = DefinedName('Duration_Hours', attr_text='Inputs!$C$9')
    wb.defined_names['Energy_MWh'] = DefinedName('Energy_MWh', attr_text='Inputs!$C$10')
    wb.defined_names['Analysis_Years'] = DefinedName('Analysis_Years', attr_text='Inputs!$C$11')
    wb.defined_names['Discount_Rate'] = DefinedName('Discount_Rate', attr_text='Inputs!$C$12')

    return benefits_start


def create_calculations_sheet(wb, styles):
    """Create the Calculations sheet with all formulas."""
    ws = wb.create_sheet("Calculations")

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 50

    row = 2
    ws['B2'] = "Calculation Engine"
    ws['B2'].font = Font(bold=True, size=14, color='1565C0')
    row = 4

    # Derived values
    ws[f'B{row}'] = "DERIVED VALUES"
    ws[f'B{row}'].style = styles['section']
    ws.merge_cells(f'B{row}:D{row}')
    row += 1

    calcs = [
        ('Capacity (kW)', '=Inputs!C8*1000', 'Convert MW to kW'),
        ('Capacity (kWh)', '=Inputs!C10*1000', 'Convert MWh to kWh'),
        ('Total CapEx ($)', '=Inputs!C24*C6', 'CapEx/kWh × kWh'),
        ('Annual Fixed O&M ($)', '=Inputs!C25*C5', 'FOM/kW × kW'),
        ('Augmentation Cost ($)', '=Inputs!C27*C6', 'Aug cost/kWh × kWh'),
        ('Decommissioning Cost ($)', '=Inputs!C28*C5', 'Decom/kW × kW'),
    ]

    for label, formula, desc in calcs:
        ws[f'B{row}'] = label
        ws[f'C{row}'] = formula
        ws[f'C{row}'].style = styles['formula']
        ws[f'C{row}'].number_format = '$#,##0'
        ws[f'D{row}'] = desc
        ws[f'D{row}'].font = Font(italic=True, color='666666')
        row += 1

    row += 1

    # Annual energy calculations
    ws[f'B{row}'] = "ANNUAL ENERGY (Year 1)"
    ws[f'B{row}'].style = styles['section']
    ws.merge_cells(f'B{row}:D{row}')
    row += 1

    ws[f'B{row}'] = "Cycles per year"
    cycles_row = row
    ws[f'C{row}'] = 365
    ws[f'C{row}'].style = styles['input']
    ws[f'D{row}'] = "Assumed 1 cycle per day"
    row += 1

    ws[f'B{row}'] = "Annual discharge (MWh)"
    ws[f'C{row}'] = f'=Inputs!C10*C{cycles_row}*Inputs!C18'
    ws[f'C{row}'].style = styles['formula']
    ws[f'C{row}'].number_format = '#,##0'
    ws[f'D{row}'] = "MWh × cycles × RTE"
    energy_row = row
    row += 2

    # NPV formula explanation
    ws[f'B{row}'] = "KEY FORMULAS"
    ws[f'B{row}'].style = styles['section']
    ws.merge_cells(f'B{row}:D{row}')
    row += 1

    formulas = [
        ('NPV', 'Σ(CFₜ / (1+r)ᵗ) for t=0 to N', 'Brealey & Myers (2020)'),
        ('BCR', 'PV(Benefits) / PV(Costs)', 'CPUC Standard Practice Manual'),
        ('IRR', 'Rate where NPV = 0', 'Brealey & Myers (2020)'),
        ('LCOS', 'PV(Costs) / PV(Energy)', 'Lazard Methodology'),
        ('Payback', 'Year where cumulative CF > 0', 'Standard finance'),
    ]

    for metric, formula, source in formulas:
        ws[f'B{row}'] = metric
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'C{row}'] = formula
        ws[f'D{row}'] = source
        ws[f'D{row}'].font = Font(italic=True)
        row += 1


def create_cashflows_sheet(wb, styles):
    """Create annual cash flow projections sheet."""
    ws = wb.create_sheet("Cash_Flows")

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 8

    # Headers
    ws['B2'] = "Annual Cash Flow Projections"
    ws['B2'].font = Font(bold=True, size=14, color='1565C0')

    row = 4
    headers = ['Year', 'CapEx', 'Fixed O&M', 'Var O&M', 'Augmentation',
               'Decommission', 'Total Costs', 'RA Benefit', 'Arbitrage',
               'Ancillary', 'T&D Defer', 'Total Benefits', 'Net Cash Flow',
               'Discount Factor', 'PV Costs', 'PV Benefits', 'Cumulative CF']

    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=row, column=col, value=header)
        cell.style = styles['header']
        ws.column_dimensions[get_column_letter(col)].width = 14

    row += 1
    start_row = row

    # Generate 21 years (0-20)
    for year in range(21):
        col = 2
        # Year
        ws.cell(row=row, column=col, value=year)
        col += 1

        # CapEx (year 0 only)
        if year == 0:
            ws.cell(row=row, column=col, value='=Calculations!C7')
        else:
            ws.cell(row=row, column=col, value=0)
        ws.cell(row=row, column=col).number_format = '$#,##0'
        col += 1

        # Fixed O&M (years 1+)
        if year == 0:
            ws.cell(row=row, column=col, value=0)
        else:
            ws.cell(row=row, column=col, value='=Calculations!C8')
        ws.cell(row=row, column=col).number_format = '$#,##0'
        col += 1

        # Variable O&M (simplified)
        ws.cell(row=row, column=col, value=0)
        ws.cell(row=row, column=col).number_format = '$#,##0'
        col += 1

        # Augmentation (year 12) - adjusted for learning curve cost decline
        if year == 12:
            # Augmentation cost = base cost × (1 - learning_rate)^years_from_base
            # Learning rate is at row 39 in Inputs sheet
            ws.cell(row=row, column=col, value='=Calculations!C9*(1-Inputs!$C$39)^12')
        else:
            ws.cell(row=row, column=col, value=0)
        ws.cell(row=row, column=col).number_format = '$#,##0'
        col += 1

        # Decommissioning (year 20)
        if year == 20:
            ws.cell(row=row, column=col, value='=Calculations!C10')
        else:
            ws.cell(row=row, column=col, value=0)
        ws.cell(row=row, column=col).number_format = '$#,##0'
        col += 1

        # Total Costs
        ws.cell(row=row, column=col, value=f'=SUM(C{row}:G{row})')
        ws.cell(row=row, column=col).number_format = '$#,##0'
        ws.cell(row=row, column=col).style = styles['formula']
        col += 1

        # Benefits (with escalation) - RA, Arbitrage, Ancillary, T&D
        if year == 0:
            for _ in range(4):
                ws.cell(row=row, column=col, value=0)
                ws.cell(row=row, column=col).number_format = '$#,##0'
                col += 1
        else:
            # RA with escalation
            esc_factor = f'*(1+Inputs!$D$32)^{year-1}'
            degradation = f'*(1-Inputs!$C$19)^{year-1}'

            ws.cell(row=row, column=col, value=f'=Inputs!$C$32*Inputs!$C$8*1000{esc_factor}{degradation}')
            ws.cell(row=row, column=col).number_format = '$#,##0'
            col += 1

            ws.cell(row=row, column=col, value=f'=Inputs!$C$33*Inputs!$C$8*1000*(1+Inputs!$D$33)^{year-1}{degradation}')
            ws.cell(row=row, column=col).number_format = '$#,##0'
            col += 1

            ws.cell(row=row, column=col, value=f'=Inputs!$C$34*Inputs!$C$8*1000*(1+Inputs!$D$34)^{year-1}{degradation}')
            ws.cell(row=row, column=col).number_format = '$#,##0'
            col += 1

            ws.cell(row=row, column=col, value=f'=Inputs!$C$35*Inputs!$C$8*1000*(1+Inputs!$D$35)^{year-1}')
            ws.cell(row=row, column=col).number_format = '$#,##0'
            col += 1

        # Total Benefits
        ws.cell(row=row, column=col, value=f'=SUM(I{row}:L{row})')
        ws.cell(row=row, column=col).number_format = '$#,##0'
        ws.cell(row=row, column=col).style = styles['formula']
        col += 1

        # Net Cash Flow
        ws.cell(row=row, column=col, value=f'=M{row}-H{row}')
        ws.cell(row=row, column=col).number_format = '$#,##0'
        ws.cell(row=row, column=col).style = styles['formula']
        col += 1

        # Discount Factor
        ws.cell(row=row, column=col, value=f'=1/(1+Inputs!$C$12)^B{row}')
        ws.cell(row=row, column=col).number_format = '0.0000'
        col += 1

        # PV Costs
        ws.cell(row=row, column=col, value=f'=H{row}*O{row}')
        ws.cell(row=row, column=col).number_format = '$#,##0'
        col += 1

        # PV Benefits
        ws.cell(row=row, column=col, value=f'=M{row}*O{row}')
        ws.cell(row=row, column=col).number_format = '$#,##0'
        col += 1

        # Cumulative CF
        if year == 0:
            ws.cell(row=row, column=col, value=f'=N{row}')
        else:
            ws.cell(row=row, column=col, value=f'=R{row-1}+N{row}')
        ws.cell(row=row, column=col).number_format = '$#,##0'

        row += 1

    end_row = row - 1

    # Totals row
    row += 1
    ws.cell(row=row, column=2, value='TOTALS')
    ws.cell(row=row, column=2).font = Font(bold=True)

    ws.cell(row=row, column=8, value=f'=SUM(H{start_row}:H{end_row})')
    ws.cell(row=row, column=8).number_format = '$#,##0'
    ws.cell(row=row, column=8).font = Font(bold=True)

    ws.cell(row=row, column=13, value=f'=SUM(M{start_row}:M{end_row})')
    ws.cell(row=row, column=13).number_format = '$#,##0'
    ws.cell(row=row, column=13).font = Font(bold=True)

    ws.cell(row=row, column=16, value=f'=SUM(P{start_row}:P{end_row})')
    ws.cell(row=row, column=16).number_format = '$#,##0'
    ws.cell(row=row, column=16).font = Font(bold=True)

    ws.cell(row=row, column=17, value=f'=SUM(Q{start_row}:Q{end_row})')
    ws.cell(row=row, column=17).number_format = '$#,##0'
    ws.cell(row=row, column=17).font = Font(bold=True)

    # Store totals row for Results sheet reference
    return start_row, end_row, row


def create_results_sheet(wb, styles, cf_totals_row):
    """Create the Results Dashboard sheet."""
    ws = wb.create_sheet("Results")

    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 35

    row = 2
    ws['B2'] = "BESS Economic Analysis Results"
    ws['B2'].font = Font(bold=True, size=18, color='1565C0')
    ws.merge_cells('B2:E2')
    row = 4

    # Project Summary
    ws[f'B{row}'] = "PROJECT SUMMARY"
    ws[f'B{row}'].style = styles['section']
    ws.merge_cells(f'B{row}:E{row}')
    row += 1

    summary = [
        ('Project', '=Inputs!C5'),
        ('Location', '=Inputs!C7'),
        ('Capacity', '=Inputs!C8&" MW / "&Inputs!C10&" MWh"'),
        ('Total Investment', '=Calculations!C7'),
    ]

    for label, formula in summary:
        ws[f'B{row}'] = label
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'C{row}'] = formula
        if 'Investment' in label:
            ws[f'C{row}'].number_format = '$#,##0'
        row += 1

    row += 1

    # Key Metrics
    ws[f'B{row}'] = "KEY FINANCIAL METRICS"
    ws[f'B{row}'].style = styles['section']
    ws.merge_cells(f'B{row}:E{row}')
    row += 1

    metrics_start = row
    metrics = [
        ('Benefit-Cost Ratio (BCR)', f'=Cash_Flows!Q{cf_totals_row}/Cash_Flows!P{cf_totals_row}', '0.00',
         'BCR ≥ 1.0 indicates benefits exceed costs'),
        ('Net Present Value (NPV)', f'=Cash_Flows!Q{cf_totals_row}-Cash_Flows!P{cf_totals_row}', '$#,##0',
         'Positive NPV indicates value creation'),
        ('Internal Rate of Return (IRR)', f'=IRR(Cash_Flows!N5:N25)', '0.0%',
         'Rate where NPV = 0'),
        ('Simple Payback (years)', f'=MATCH(TRUE,Cash_Flows!R5:R25>0,0)-1', '0.0',
         'Year when cumulative CF turns positive'),
        ('PV of Total Costs', f'=Cash_Flows!P{cf_totals_row}', '$#,##0,,"M"',
         'Present value of all costs'),
        ('PV of Total Benefits', f'=Cash_Flows!Q{cf_totals_row}', '$#,##0,,"M"',
         'Present value of all benefits'),
    ]

    for label, formula, fmt, desc in metrics:
        ws[f'B{row}'] = label
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'C{row}'] = formula
        ws[f'C{row}'].number_format = fmt
        ws[f'C{row}'].style = styles['result']
        ws[f'E{row}'] = desc
        ws[f'E{row}'].font = Font(italic=True, color='666666')
        row += 1

    # Add conditional formatting to BCR
    bcr_cell = f'C{metrics_start}'
    ws.conditional_formatting.add(
        bcr_cell,
        FormulaRule(formula=[f'{bcr_cell}>=1.5'], fill=PatternFill('solid', fgColor='C8E6C9'))
    )
    ws.conditional_formatting.add(
        bcr_cell,
        FormulaRule(formula=[f'AND({bcr_cell}>=1,{bcr_cell}<1.5)'], fill=PatternFill('solid', fgColor='FFF9C4'))
    )
    ws.conditional_formatting.add(
        bcr_cell,
        FormulaRule(formula=[f'{bcr_cell}<1'], fill=PatternFill('solid', fgColor='FFCDD2'))
    )

    row += 2

    # Recommendation
    ws[f'B{row}'] = "RECOMMENDATION"
    ws[f'B{row}'].style = styles['section']
    ws.merge_cells(f'B{row}:E{row}')
    row += 1

    ws[f'B{row}'] = '=IF(C11>=1.5,"APPROVE - Strong economic case",IF(C11>=1,"FURTHER STUDY - Marginal economics","REJECT - Costs exceed benefits"))'
    ws.merge_cells(f'B{row}:E{row}')
    ws[f'B{row}'].font = Font(bold=True, size=14)
    row += 2

    # LCOS Calculation
    ws[f'B{row}'] = "LEVELIZED COST OF STORAGE (LCOS)"
    ws[f'B{row}'].style = styles['section']
    ws.merge_cells(f'B{row}:E{row}')
    row += 1

    ws[f'B{row}'] = "LCOS ($/MWh)"
    ws[f'B{row}'].font = Font(bold=True)
    # LCOS = PV Costs / PV Energy Discharged
    ws[f'C{row}'] = f'=Cash_Flows!P{cf_totals_row}/(Calculations!C13*NPV(Inputs!C12,INDIRECT("Cash_Flows!O5:O25")))'
    ws[f'C{row}'].number_format = '$#,##0.00'
    ws[f'C{row}'].style = styles['result']
    ws[f'E{row}'] = "Levelized cost per MWh discharged (Lazard methodology)"
    row += 2

    # Breakeven Analysis
    ws[f'B{row}'] = "BREAKEVEN ANALYSIS"
    ws[f'B{row}'].style = styles['section']
    ws.merge_cells(f'B{row}:E{row}')
    row += 1

    ws[f'B{row}'] = "Breakeven CapEx ($/kWh)"
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'C{row}'] = f'=(Cash_Flows!Q{cf_totals_row}-(Cash_Flows!P{cf_totals_row}-Calculations!C7))/(Inputs!C10*1000)'
    ws[f'C{row}'].number_format = '$#,##0'
    ws[f'C{row}'].style = styles['result']
    ws[f'E{row}'] = "Maximum CapEx for BCR = 1.0"

    return metrics_start


def create_methodology_sheet(wb, styles):
    """Create methodology documentation sheet."""
    ws = wb.create_sheet("Methodology")

    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 80

    row = 2
    ws['B2'] = "Methodology & References"
    ws['B2'].font = Font(bold=True, size=16, color='1565C0')
    row = 4

    # Overview
    ws[f'B{row}'] = "OVERVIEW"
    ws[f'B{row}'].style = styles['section']
    row += 1

    overview = """This workbook performs economic analysis of battery energy storage system (BESS) projects using standard discounted cash flow (DCF) methodology. All calculations follow industry-standard approaches from NREL, Lazard, and the California Public Utilities Commission (CPUC)."""
    ws[f'B{row}'] = overview
    ws[f'B{row}'].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[row].height = 45
    row += 2

    # Formulas Section
    ws[f'B{row}'] = "KEY FORMULAS"
    ws[f'B{row}'].style = styles['section']
    row += 1

    formulas = [
        ("Net Present Value (NPV)",
         "NPV = Σ(CFₜ / (1+r)ᵗ) for t=0 to N",
         "Where CFₜ = cash flow in year t, r = discount rate, N = analysis period"),

        ("Benefit-Cost Ratio (BCR)",
         "BCR = PV(Benefits) / PV(Costs)",
         "BCR > 1.0 indicates project creates net value; BCR > 1.5 is strong"),

        ("Internal Rate of Return (IRR)",
         "IRR: Find r where NPV = 0",
         "The discount rate that makes NPV equal to zero"),

        ("Levelized Cost of Storage (LCOS)",
         "LCOS = PV(Lifetime Costs) / PV(Lifetime Energy Discharged)",
         "Expressed in $/MWh; enables comparison across technologies"),

        ("Payback Period",
         "Year t where Σ(CF₀ to CFₜ) ≥ 0",
         "First year when cumulative cash flow turns positive"),

        ("Degradation Adjustment",
         "Capacity(t) = Capacity(0) × (1 - degradation_rate)^t",
         "Applied to capacity-based revenue streams"),

        ("Benefit Escalation",
         "Value(t) = Value(1) × (1 + escalation_rate)^(t-1)",
         "Annual price escalation for each benefit stream"),

        ("Technology Learning Curve (Cost Decline)",
         "Cost(year) = Cost(base) × (1 - learning_rate)^(year - base_year)",
         "Battery costs decline ~10-15% annually; reduces augmentation & replacement costs"),

        ("T&D Deferral Benefit",
         "Annual Value = Avoided_Cost_per_kW × Capacity_kW × BESS_Contribution_Factor",
         "Value of deferring transmission/distribution infrastructure investments"),
    ]

    for name, formula, description in formulas:
        ws[f'B{row}'] = name
        ws[f'B{row}'].font = Font(bold=True)
        row += 1
        ws[f'B{row}'] = f"Formula: {formula}"
        ws[f'B{row}'].font = Font(italic=True)
        row += 1
        ws[f'B{row}'] = description
        row += 2

    # Assumption Libraries
    ws[f'B{row}'] = "ASSUMPTION LIBRARIES"
    ws[f'B{row}'].style = styles['section']
    row += 1

    libraries = [
        ("NREL ATB 2024 - Moderate",
         "National Renewable Energy Laboratory Annual Technology Baseline 2024",
         "https://atb.nrel.gov/electricity/2024/utility-scale_battery_storage",
         "Moderate cost scenario from median of 16 industry sources. 4-hour LFP system."),

        ("Lazard LCOS 2025",
         "Lazard's Levelized Cost of Storage Analysis, Version 10.0 (March 2025)",
         "https://www.lazard.com/research-insights/levelized-cost-of-storage/",
         "Utility-scale 100 MW / 400 MWh LFP system. Merchant project structure."),

        ("CPUC California 2024",
         "California Public Utilities Commission Standard Practice Manual",
         "https://www.cpuc.ca.gov/industries-and-topics/electrical-energy/demand-side-management",
         "California-specific assumptions including RA premiums and T&D deferral values."),
    ]

    for name, source, url, notes in libraries:
        ws[f'B{row}'] = name
        ws[f'B{row}'].font = Font(bold=True)
        row += 1
        ws[f'B{row}'] = f"Source: {source}"
        row += 1
        ws[f'B{row}'] = f"URL: {url}"
        ws[f'B{row}'].font = Font(color='0066CC', underline='single')
        row += 1
        ws[f'B{row}'] = f"Notes: {notes}"
        row += 2

    # Full Citations
    ws[f'B{row}'] = "REFERENCES & CITATIONS"
    ws[f'B{row}'].style = styles['section']
    row += 1

    citations = [
        "[1] NREL. Annual Technology Baseline 2024. National Renewable Energy Laboratory. April 2024. https://atb.nrel.gov/",

        "[2] Lazard. Lazard's Levelized Cost of Storage Analysis, Version 10.0. March 2025.",

        "[3] California Public Utilities Commission. California Standard Practice Manual: Economic Analysis of Demand-Side Programs and Projects. October 2001.",

        "[4] California Public Utilities Commission. 2024 Resource Adequacy Report. D.24-06-050. November 2024.",

        "[5] E3 (Energy+Environmental Economics). 2024 Avoided Cost Calculator. Prepared for CPUC. October 2024.",

        "[6] California ISO. Market Performance Reports 2024. https://www.caiso.com/market/",

        "[7] Brealey, R., Myers, S., & Allen, F. Principles of Corporate Finance (13th ed.). McGraw-Hill, 2020.",

        "[8] NREL. Storage Futures Study: Economic Potential of Diurnal Storage. NREL/TP-6A20-77449. 2021.",

        "[9] BloombergNEF. Lithium-Ion Battery Price Survey 2025. Battery technology learning curves.",

        "[10] E3. California Energy Storage Cost Analysis 2024. T&D Deferral methodology and avoided cost calculations.",
    ]

    for cite in citations:
        ws[f'B{row}'] = cite
        ws[f'B{row}'].alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row].height = 30
        row += 1

    row += 1
    ws[f'B{row}'] = "DISCLAIMER"
    ws[f'B{row}'].style = styles['section']
    row += 1

    disclaimer = """This model is provided for analytical purposes only. Results should be validated against independent sources before use in investment decisions. Actual project economics will vary based on site-specific conditions, market dynamics, and evolving technology costs. Users are responsible for verifying all assumptions and calculations."""
    ws[f'B{row}'] = disclaimer
    ws[f'B{row}'].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[row].height = 60


def create_libraries_sheet(wb, styles):
    """Create sheet with all assumption library data for VLOOKUP reference."""
    ws = wb.create_sheet("Library_Data")

    ws.column_dimensions['A'].width = 30
    for col in 'BCDEFGH':
        ws.column_dimensions[col].width = 15

    row = 1
    ws['A1'] = "Assumption Library Data (for reference)"
    ws['A1'].font = Font(bold=True, size=12)
    row = 3

    # Cost parameters
    headers = ['Parameter', 'NREL ATB 2024', 'Lazard LCOS 2025', 'CPUC California 2024']
    for col, header in enumerate(headers, start=1):
        ws.cell(row=row, column=col, value=header)
        ws.cell(row=row, column=col).style = styles['header']
    row += 1

    cost_data = [
        ('CapEx ($/kWh)', 160, 145, 155),
        ('Fixed O&M ($/kW-yr)', 25, 22, 26),
        ('Variable O&M ($/MWh)', 0, 0.5, 0),
        ('Augmentation ($/kWh)', 55, 50, 52),
        ('Decommissioning ($/kW)', 10, 8, 12),
        ('Round-Trip Efficiency', '85%', '86%', '85%'),
        ('Annual Degradation', '2.5%', '2.0%', '2.5%'),
        ('Cycle Life', 6000, 6500, 6000),
        ('Augmentation Year', 12, 12, 12),
        ('Learning Rate (Cost Decline)', '12%', '10%', '11%'),
        ('Cost Base Year', 2024, 2025, 2024),
    ]

    for param, nrel, lazard, cpuc in cost_data:
        ws.cell(row=row, column=1, value=param)
        ws.cell(row=row, column=2, value=nrel)
        ws.cell(row=row, column=3, value=lazard)
        ws.cell(row=row, column=4, value=cpuc)
        row += 1

    row += 1

    # Tax Credits (BESS-Specific)
    ws.cell(row=row, column=1, value="Tax Credits (BESS-Specific)")
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1

    tax_data = [
        ('ITC Base Rate', '30%', '30%', '30%'),
        ('ITC Adders', '0%', '0%', '10%'),
    ]

    for param, nrel, lazard, cpuc in tax_data:
        ws.cell(row=row, column=1, value=param)
        ws.cell(row=row, column=2, value=nrel)
        ws.cell(row=row, column=3, value=lazard)
        ws.cell(row=row, column=4, value=cpuc)
        row += 1

    row += 1

    # Infrastructure Costs (Common)
    ws.cell(row=row, column=1, value="Infrastructure Costs (Common)")
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1

    infra_data = [
        ('Interconnection ($/kW)', 100, 90, 120),
        ('Land ($/kW)', 10, 8, 15),
        ('Permitting ($/kW)', 15, 12, 20),
        ('Insurance (% of CapEx)', '0.5%', '0.5%', '0.5%'),
        ('Property Tax (%)', '1.0%', '1.0%', '1.05%'),
    ]

    for param, nrel, lazard, cpuc in infra_data:
        ws.cell(row=row, column=1, value=param)
        ws.cell(row=row, column=2, value=nrel)
        ws.cell(row=row, column=3, value=lazard)
        ws.cell(row=row, column=4, value=cpuc)
        row += 1

    row += 1

    # Benefit parameters
    ws.cell(row=row, column=1, value="Benefit Streams ($/kW-yr)")
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1

    benefit_data = [
        ('Resource Adequacy [common]', 150, 140, 180),
        ('Energy Arbitrage [common]', 40, 45, 35),
        ('Ancillary Services [common]', 15, 12, 10),
        ('T&D Deferral [common]', 25, 20, 25),
        ('Resilience Value [common]', 50, 45, 60),
        ('Renewable Integration [bess]', 25, 20, 30),
        ('GHG Emissions Value [bess]', 15, 12, 20),
        ('Voltage Support [common]', 8, 6, 10),
    ]

    for param, nrel, lazard, cpuc in benefit_data:
        ws.cell(row=row, column=1, value=param)
        ws.cell(row=row, column=2, value=nrel)
        ws.cell(row=row, column=3, value=lazard)
        ws.cell(row=row, column=4, value=cpuc)
        for col in [2, 3, 4]:
            ws.cell(row=row, column=col).number_format = '$#,##0'
        row += 1


def create_vba_instructions_sheet(wb, styles):
    """Create sheet with VBA macro instructions."""
    ws = wb.create_sheet("VBA_Instructions")

    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 100

    row = 2
    ws['B2'] = "VBA Macros & Report Generation"
    ws['B2'].font = Font(bold=True, size=16, color='1565C0')
    row = 4

    ws[f'B{row}'] = "IMPORTANT: TO ENABLE MACROS"
    ws[f'B{row}'].style = styles['section']
    row += 1

    instructions = [
        "1. Save this file as .xlsm (Excel Macro-Enabled Workbook)",
        "2. Open the VBA Editor (Alt+F11 on Windows, Option+F11 on Mac)",
        "3. Insert a new Module (Insert > Module)",
        "4. Copy and paste the VBA code below into the module",
        "5. Save and close the VBA Editor",
        "6. You can now run macros from the Developer tab or assign to buttons",
    ]

    for instr in instructions:
        ws[f'B{row}'] = instr
        row += 1

    row += 1
    ws[f'B{row}'] = "VBA CODE TO COPY:"
    ws[f'B{row}'].style = styles['section']
    row += 1

    vba_code = '''
'=============================================================
' BESS Analyzer VBA Macros
' Copy this entire code block into a VBA Module
'=============================================================

Sub GenerateReport()
    ' Generate a formatted report on a new sheet
    Dim wsReport As Worksheet
    Dim wsResults As Worksheet
    Dim wsInputs As Worksheet
    Dim wsCF As Worksheet

    Set wsResults = ThisWorkbook.Sheets("Results")
    Set wsInputs = ThisWorkbook.Sheets("Inputs")
    Set wsCF = ThisWorkbook.Sheets("Cash_Flows")

    ' Delete old report if exists
    On Error Resume Next
    Application.DisplayAlerts = False
    ThisWorkbook.Sheets("Report").Delete
    Application.DisplayAlerts = True
    On Error GoTo 0

    ' Create new report sheet
    Set wsReport = ThisWorkbook.Sheets.Add(After:=ThisWorkbook.Sheets(ThisWorkbook.Sheets.Count))
    wsReport.Name = "Report"

    ' Set up report formatting
    wsReport.PageSetup.Orientation = xlPortrait
    wsReport.PageSetup.FitToPagesWide = 1
    wsReport.PageSetup.FitToPagesTall = False

    Dim row As Long
    row = 2

    ' Title
    wsReport.Cells(row, 2).Value = "BESS ECONOMIC ANALYSIS REPORT"
    wsReport.Cells(row, 2).Font.Bold = True
    wsReport.Cells(row, 2).Font.Size = 18
    wsReport.Range("B" & row & ":F" & row).Merge
    row = row + 2

    ' Project Summary
    wsReport.Cells(row, 2).Value = "PROJECT SUMMARY"
    wsReport.Cells(row, 2).Font.Bold = True
    wsReport.Cells(row, 2).Font.Size = 14
    wsReport.Cells(row, 2).Interior.Color = RGB(227, 242, 253)
    wsReport.Range("B" & row & ":F" & row).Merge
    row = row + 1

    wsReport.Cells(row, 2).Value = "Project Name:"
    wsReport.Cells(row, 3).Value = wsInputs.Range("C5").Value
    row = row + 1

    wsReport.Cells(row, 2).Value = "Location:"
    wsReport.Cells(row, 3).Value = wsInputs.Range("C7").Value
    row = row + 1

    wsReport.Cells(row, 2).Value = "Capacity:"
    wsReport.Cells(row, 3).Value = wsInputs.Range("C8").Value & " MW / " & wsInputs.Range("C10").Value & " MWh"
    row = row + 1

    wsReport.Cells(row, 2).Value = "Total Investment:"
    wsReport.Cells(row, 3).Value = wsResults.Range("C8").Value
    wsReport.Cells(row, 3).NumberFormat = "$#,##0"
    row = row + 2

    ' Key Metrics
    wsReport.Cells(row, 2).Value = "KEY FINANCIAL METRICS"
    wsReport.Cells(row, 2).Font.Bold = True
    wsReport.Cells(row, 2).Font.Size = 14
    wsReport.Cells(row, 2).Interior.Color = RGB(227, 242, 253)
    wsReport.Range("B" & row & ":F" & row).Merge
    row = row + 1

    wsReport.Cells(row, 2).Value = "Benefit-Cost Ratio (BCR):"
    wsReport.Cells(row, 3).Value = wsResults.Range("C11").Value
    wsReport.Cells(row, 3).NumberFormat = "0.00"
    row = row + 1

    wsReport.Cells(row, 2).Value = "Net Present Value (NPV):"
    wsReport.Cells(row, 3).Value = wsResults.Range("C12").Value
    wsReport.Cells(row, 3).NumberFormat = "$#,##0"
    row = row + 1

    wsReport.Cells(row, 2).Value = "Internal Rate of Return (IRR):"
    wsReport.Cells(row, 3).Value = wsResults.Range("C13").Value
    wsReport.Cells(row, 3).NumberFormat = "0.0%"
    row = row + 1

    wsReport.Cells(row, 2).Value = "Simple Payback Period:"
    wsReport.Cells(row, 3).Value = wsResults.Range("C14").Value & " years"
    row = row + 1

    wsReport.Cells(row, 2).Value = "LCOS:"
    wsReport.Cells(row, 3).Value = wsResults.Range("C21").Value
    wsReport.Cells(row, 3).NumberFormat = "$#,##0.00/MWh"
    row = row + 2

    ' Recommendation
    wsReport.Cells(row, 2).Value = "RECOMMENDATION"
    wsReport.Cells(row, 2).Font.Bold = True
    wsReport.Cells(row, 2).Font.Size = 14
    wsReport.Cells(row, 2).Interior.Color = RGB(227, 242, 253)
    wsReport.Range("B" & row & ":F" & row).Merge
    row = row + 1

    wsReport.Cells(row, 2).Value = wsResults.Range("B19").Value
    wsReport.Cells(row, 2).Font.Bold = True
    wsReport.Cells(row, 2).Font.Size = 12
    wsReport.Range("B" & row & ":F" & row).Merge
    row = row + 2

    ' Methodology Note
    wsReport.Cells(row, 2).Value = "METHODOLOGY"
    wsReport.Cells(row, 2).Font.Bold = True
    wsReport.Cells(row, 2).Font.Size = 14
    wsReport.Cells(row, 2).Interior.Color = RGB(227, 242, 253)
    wsReport.Range("B" & row & ":F" & row).Merge
    row = row + 1

    wsReport.Cells(row, 2).Value = "Analysis uses standard DCF methodology per CPUC Standard Practice Manual."
    row = row + 1
    wsReport.Cells(row, 2).Value = "BCR = PV(Benefits) / PV(Costs); IRR = rate where NPV = 0"
    row = row + 1
    wsReport.Cells(row, 2).Value = "LCOS calculated per Lazard methodology: PV(Costs) / PV(Energy)"
    row = row + 2

    ' References
    wsReport.Cells(row, 2).Value = "REFERENCES"
    wsReport.Cells(row, 2).Font.Bold = True
    wsReport.Cells(row, 2).Font.Size = 14
    wsReport.Cells(row, 2).Interior.Color = RGB(227, 242, 253)
    wsReport.Range("B" & row & ":F" & row).Merge
    row = row + 1

    wsReport.Cells(row, 2).Value = "1. NREL Annual Technology Baseline 2024"
    row = row + 1
    wsReport.Cells(row, 2).Value = "2. Lazard LCOS v10.0 (2025)"
    row = row + 1
    wsReport.Cells(row, 2).Value = "3. CPUC Standard Practice Manual (2001)"
    row = row + 1
    wsReport.Cells(row, 2).Value = "4. Brealey & Myers, Principles of Corporate Finance"
    row = row + 2

    ' Adjust column widths
    wsReport.Columns("B").ColumnWidth = 30
    wsReport.Columns("C").ColumnWidth = 25
    wsReport.Columns("D:F").ColumnWidth = 15

    MsgBox "Report generated on 'Report' sheet. You can print or export to PDF.", vbInformation

End Sub

Sub ExportToPDF()
    ' Export the Report sheet to PDF
    Dim wsReport As Worksheet
    Dim filePath As String

    On Error Resume Next
    Set wsReport = ThisWorkbook.Sheets("Report")
    On Error GoTo 0

    If wsReport Is Nothing Then
        MsgBox "Please run 'Generate Report' first.", vbExclamation
        Exit Sub
    End If

    filePath = Application.GetSaveAsFilename( _
        InitialFileName:="BESS_Analysis_Report.pdf", _
        FileFilter:="PDF Files (*.pdf), *.pdf")

    If filePath <> "False" Then
        wsReport.ExportAsFixedFormat Type:=xlTypePDF, Filename:=filePath
        MsgBox "PDF exported to: " & filePath, vbInformation
    End If

End Sub

Sub RefreshCalculations()
    ' Force recalculation of all formulas
    Application.CalculateFull
    MsgBox "All calculations refreshed.", vbInformation
End Sub

Sub LoadNRELLibrary()
    ' Load NREL ATB 2024 assumptions
    With ThisWorkbook.Sheets("Inputs")
        .Range("C14").Value = "NREL ATB 2024 - Moderate"
        .Range("C17").Value = "LFP"
        .Range("C18").Value = 0.85
        .Range("C19").Value = 0.025
        .Range("C20").Value = 6000
        .Range("C21").Value = 12
        .Range("C24").Value = 160
        .Range("C25").Value = 25
        .Range("C26").Value = 0
        .Range("C27").Value = 55
        .Range("C28").Value = 10
        ' Benefit streams
        .Range("C32").Value = 150
        .Range("D32").Value = 0.02
        .Range("C33").Value = 40
        .Range("D33").Value = 0.015
        .Range("C34").Value = 15
        .Range("D34").Value = 0.01
        .Range("C35").Value = 25   ' T&D Deferral
        .Range("D35").Value = 0.02
        ' Cost projections (learning curve)
        .Range("C38").Value = 0.12  ' 12% annual cost decline
        .Range("C39").Value = 2024  ' Base year
    End With
    MsgBox "NREL ATB 2024 assumptions loaded." & vbCrLf & _
           "Includes 12% annual cost decline for future augmentation.", vbInformation
End Sub

Sub LoadLazardLibrary()
    ' Load Lazard LCOS 2025 assumptions
    With ThisWorkbook.Sheets("Inputs")
        .Range("C14").Value = "Lazard LCOS 2025"
        .Range("C17").Value = "LFP"
        .Range("C18").Value = 0.86
        .Range("C19").Value = 0.02
        .Range("C20").Value = 6500
        .Range("C21").Value = 12
        .Range("C24").Value = 145
        .Range("C25").Value = 22
        .Range("C26").Value = 0.5
        .Range("C27").Value = 50
        .Range("C28").Value = 8
        ' Benefit streams
        .Range("C32").Value = 140
        .Range("D32").Value = 0.02
        .Range("C33").Value = 45
        .Range("D33").Value = 0.02
        .Range("C34").Value = 12
        .Range("D34").Value = 0.01
        .Range("C35").Value = 20   ' T&D Deferral
        .Range("D35").Value = 0.015
        ' Cost projections (learning curve)
        .Range("C38").Value = 0.10  ' 10% annual cost decline
        .Range("C39").Value = 2025  ' Base year
    End With
    MsgBox "Lazard LCOS 2025 assumptions loaded." & vbCrLf & _
           "Includes 10% annual cost decline for future augmentation.", vbInformation
End Sub

Sub LoadCPUCLibrary()
    ' Load CPUC California 2024 assumptions
    With ThisWorkbook.Sheets("Inputs")
        .Range("C14").Value = "CPUC California 2024"
        .Range("C17").Value = "LFP"
        .Range("C18").Value = 0.85
        .Range("C19").Value = 0.025
        .Range("C20").Value = 6000
        .Range("C21").Value = 12
        .Range("C24").Value = 155
        .Range("C25").Value = 26
        .Range("C26").Value = 0
        .Range("C27").Value = 52
        .Range("C28").Value = 12
        ' Benefit streams
        .Range("C32").Value = 180
        .Range("D32").Value = 0.025
        .Range("C33").Value = 35
        .Range("D33").Value = 0.02
        .Range("C34").Value = 10
        .Range("D34").Value = 0.01
        .Range("C35").Value = 25   ' T&D Deferral
        .Range("D35").Value = 0.015
        ' Cost projections (learning curve)
        .Range("C38").Value = 0.11  ' 11% annual cost decline
        .Range("C39").Value = 2024  ' Base year
    End With
    MsgBox "CPUC California 2024 assumptions loaded." & vbCrLf & _
           "Includes 11% annual cost decline for future augmentation.", vbInformation
End Sub
'''

    for line in vba_code.strip().split('\n'):
        ws[f'B{row}'] = line
        ws[f'B{row}'].font = Font(name='Courier New', size=9)
        row += 1

    row += 2
    ws[f'B{row}'] = "ASSIGNING MACROS TO BUTTONS:"
    ws[f'B{row}'].style = styles['section']
    row += 1

    button_instructions = [
        "1. Go to Developer tab > Insert > Button (Form Control)",
        "2. Draw the button on the Inputs sheet",
        "3. When prompted, select the macro to assign (e.g., 'LoadNRELLibrary')",
        "4. Right-click button to edit text (e.g., 'Load NREL Library')",
        "",
        "Recommended buttons to create:",
        "  - 'Load NREL Library' -> LoadNRELLibrary",
        "  - 'Load Lazard Library' -> LoadLazardLibrary",
        "  - 'Load CPUC Library' -> LoadCPUCLibrary",
        "  - 'Generate Report' -> GenerateReport",
        "  - 'Export to PDF' -> ExportToPDF",
    ]

    for instr in button_instructions:
        ws[f'B{row}'] = instr
        row += 1


def create_workbook(output_path: str):
    """Create the complete BESS Analyzer workbook."""
    wb = Workbook()
    styles = create_styles()

    # Register styles
    for style in styles.values():
        if isinstance(style, NamedStyle):
            try:
                wb.add_named_style(style)
            except ValueError:
                pass

    # Create sheets
    benefits_start = create_inputs_sheet(wb, styles)
    create_calculations_sheet(wb, styles)
    cf_start, cf_end, cf_totals = create_cashflows_sheet(wb, styles)
    create_results_sheet(wb, styles, cf_totals)
    create_methodology_sheet(wb, styles)
    create_libraries_sheet(wb, styles)
    create_vba_instructions_sheet(wb, styles)

    # Reorder sheets
    sheet_order = ['Inputs', 'Results', 'Cash_Flows', 'Calculations', 'Methodology', 'Library_Data', 'VBA_Instructions']
    for i, name in enumerate(sheet_order):
        wb.move_sheet(wb[name], offset=i - wb.sheetnames.index(name))

    # Save workbook
    wb.save(output_path)
    print(f"Workbook created: {output_path}")
    print("\nTo enable macros:")
    print("1. Open in Excel and save as .xlsm (Macro-Enabled Workbook)")
    print("2. See 'VBA_Instructions' sheet for macro code")


if __name__ == "__main__":
    output = sys.argv[1] if len(sys.argv) > 1 else "BESS_Analyzer.xlsx"
    create_workbook(output)
